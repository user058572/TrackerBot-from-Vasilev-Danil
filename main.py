from typing import List, Any

import telebot
import sqlite3
from telebot import types
from time import sleep

import openpyxl
import threading

import xlsxwriter

import matplotlib.pyplot as plt
import matplotlib

matplotlib.use('agg')

from Dicts import main2, olympSheet

main_dict = {}

try:  # обновления служебных списков и словарей каждые 24 часа
    main2()
    olympSheet()

except Exception:
    flag = 1
    while flag:
        try:
            main2()
            olympSheet()
            flag = 0
        except Exception:
            print("Ошибка, повтор подключения к Google Sheet API")
            pass

token = '5846254841:AAE9hO3V9sbtkOYZeODrCdoYKvVB1FscC1I'  # Токен бота
bot = telebot.TeleBot(token)

from Dicts import lessons_dict, mentors_dict, participation_stage_dict, olympiad_dict, olympiad_lst

lessons = []

markup = 0
stickerId = 0

user_chat_id = 0
FLAG = 0
class_name = "timur_ne_viigraet_visshuy_probu"
timer = 60 * 60 * 24


def f(f_stop):  # обновление данных каждые 24 часа
    if FLAG:
        try:
            main2()
            olympSheet()
        except Exception:
            flag = 1
            while flag:
                try:
                    main2()
                    olympSheet()
                    flag = 0
                except Exception:
                    print("Ошибка, повтор подключения к Google Sheet API")
                    pass

        bot.send_message(user_chat_id,
                         "ВНИМАНИЕ! Произошло обновление базы данных. Для продолжения введите команду /start, в противном случае могут возникнуть ошибки.")
    if not f_stop.is_set():
        threading.Timer(timer, f, [f_stop]).start()


f_stop = threading.Event()
f(f_stop)


def main():
    db = sqlite3.connect("Data_Bases//Users.db",
                         check_same_thread=False)  # Создаём бд с id чатов и какие имена к ним привязаны
    cursor = db.cursor()

    db_olympiad = sqlite3.connect("Data_Bases//RsOS.db",
                                  check_same_thread=False)  # Создаём бд с именами пользователей,
    # олимпиадой, которую он написал, предметом по которой писал олимпиаду, этап олимпиады и его наставником
    cursor_olymp = db_olympiad.cursor()

    @bot.message_handler(commands=['start'])  # Обрабатываем команду start
    def start_message(message):
        global user_chat_id, FLAG
        FLAG = 1
        user_chat_id = message.from_user.id
        cursor.execute(f"SELECT chat_id FROM user WHERE chat_id = '{message.chat.id}'")
        # Ищем id чата с которого нам написали

        if cursor.fetchone() is None:  # Если такой человек нам не писал, то к id его чата не привязано имя,
            # а значит мы не сможем заносить в бд его по имени,
            # тогда нам надо спросить его имя и привязать его к id его чата
            bot.send_message(message.chat.id,
                             "✌Привет\n"
                             "Я бот для заполнения РсОШ трекера\n\n"
                             "🔎Для начала надо зарегистрироваться")  # Основное приветствие
            bot.send_message(message.chat.id, "Напиши своё ФИО и класс в формате:\n"
                                              "'reg, *твоё ФИО*, *твой класс*'\n\n"
                                              "Пример: 'reg, Иванов Иван Иванович, 11A'\n\n"
                                              "P.S. Литера латиницей! 😮 ")  # Запрос ФИО

        else:  # Если id чата уже есть в бд, значит пользователь уже зарегистрирован, тогда
            bot.send_message(message.chat.id, "Привет, ты уже зарегистрирован")  # Приветствуем его и
            button_message(message)  # Выводим ему основное меню кнопок

    def name(message):  # Функция для регистрации пользователя
        text = message.text[4:]  # Получаем ФИО пользователя
        clases = ["6a", "6b", "6c", "7a", "7b", "7c", "8a", "8b", "8c"
                                                                  "9a", "9b", "9c", "10a", "10b", "10c", "11a", "11b",
                  "11c"]
        try:
            text = text.split(", ")

            fio = text[0].rstrip().lstrip()
            clas = text[1].lower().rstrip().lstrip()

            if clas in clases:
                cursor.execute(
                    f"SELECT user_name FROM user WHERE user_name = '{fio}'")  # Проверяем, есть ли написанное имя в бд
                if cursor.fetchone() is None:  # Если его нет, тогда идём дальше
                    cursor.execute('INSERT INTO user VALUES (?, ?, ?, ?, ?)',
                                   (None, message.chat.id, fio, 0, clas))
                    # Вписываем id чата пользователя и ФИО которое он написал
                    db.commit()  # Сохраняем изменения в бд

                    bot.send_message(message.chat.id,
                                     "✅Ты успешно зарегистрировался")  # Оповещаем пользователя, что регистрация прошла успешно
                    cursor_olymp.execute("""INSERT INTO olympiad_rating VALUES (?, ?, ?)""", (None, fio, 0))
                    db_olympiad.commit()
                    button_message(message)  # Выводим основное меню с кнопками
                else:  # Если имя уже есть, тогда
                    bot.send_message(message.chat.id,
                                     "❗️Пользователь с таким именем уже зарегистрирован\n\nНажмите на /start и "
                                     "попробуйте заново")  # Говорим, что такое имя уже есть
            else:
                bot.send_message(message.chat.id,
                                 "Нарушен формат ввода класса❗\n\nНажмите на /start и "
                                 "попробуйте заново")

        except Exception:

            bot.send_message(message.chat.id,
                             "Нарушен формат ввода❗\n\nНажмите на /start и "
                             "попробуйте заново")

    def button_message(message):  # Функция для вывода основного меню кнопок
        global main_markup
        main_markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)  # Создаём меню

        my_profile_menu = types.KeyboardButton("👤Мой профиль")
        global_menu = types.KeyboardButton("📋РсОШ трекер")
        data_menu = types.KeyboardButton("📊Получить данные")

        # Создаём основные кнопки и добавляем кнопки в меню
        main_markup.add(my_profile_menu, global_menu, data_menu)
        bot.send_message(message.chat.id, '❔Что тебя интересует?',
                         reply_markup=main_markup)

    def button_menu_profile(message):
        profile_markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)

        del_olympiad = types.KeyboardButton("🗑Удалить олимпиаду")
        my_olymp = types.KeyboardButton("📄Мои олимпиады")
        rename_user = types.KeyboardButton("🔃Поменять имя и/или клас")
        back = types.KeyboardButton("🔙Назад")

        profile_markup.add(my_olymp, rename_user, del_olympiad, back)
        bot.send_message(message.chat.id, '❔Что тебя интересует?',
                         reply_markup=profile_markup)

    def button_menu_global(message):
        global_menu_markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)

        new_olymp_register = types.KeyboardButton("✏️Зарегистрироваться на олимпиаду")
        olymp_statick = types.KeyboardButton("🏆Рейтинг")
        my_place = types.KeyboardButton("📊Статистика олимпиад")
        back = types.KeyboardButton("🔙Назад")

        global_menu_markup.add(new_olymp_register, olymp_statick, my_place, back)
        bot.send_message(message.chat.id, '❔Что тебя интересует?',
                         reply_markup=global_menu_markup)

    def button_menu_data(message):
        data_menu_markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)

        table = types.KeyboardButton("👀Получить Excel-таблицу")
        back = types.KeyboardButton("🔙Назад")

        data_menu_markup.add(table, back)
        bot.send_message(message.chat.id, '❔Что тебя интересует?',
                         reply_markup=data_menu_markup)

    def receive_user_name(message):
        cursor.execute(f"SELECT user_name FROM user WHERE chat_id = '{message.chat.id}'")
        user_name = cursor.fetchone()[0]
        return user_name

    def send_olympiad_list(message, ind):  # Функция для вывода списка олимпиад
        global stickerId
        stickerId = bot.send_message(message.chat.id, "👇👇👇", reply_markup=types.ReplyKeyboardRemove())

        keyboard = types.InlineKeyboardMarkup()
        next_15 = types.InlineKeyboardButton(text="➡️Следующие 15", callback_data='next_15')
        past_15 = types.InlineKeyboardButton(text="⬅️Прошлые 15", callback_data='past_15')
        pick_olympiad = types.InlineKeyboardButton(text="✍️Выбрать олимпиаду", callback_data='pick_olympiad')

        # print(ind, len(olympiad_lst) - 1)
        if len(olympiad_lst) == 1:
            pass
        elif ind != len(olympiad_lst) - 1 and ind != 0:
            keyboard.add(past_15, next_15)
        elif ind == 0:
            keyboard.add(next_15)
        elif ind == len(olympiad_lst) - 1:
            keyboard.add(past_15)

        keyboard.add(pick_olympiad)

        try:
            bot.send_message(message.chat.id, olympiad_lst[ind], reply_markup=keyboard)
        except Exception:
            pass

    def pick_olymp(message):
        name_user = receive_user_name(message)
        try:
            olympiad_name = olympiad_dict[message.text][0]

            if len(message.text.split('_')) == 1:
                main_dict[message.chat.id] = [name_user, olympiad_name]
                pick_lesson(message)
            else:
                main_dict[message.chat.id] = [name_user, olympiad_name]
                pick_participation_stage(message)
        except KeyError:

            bot.send_message(message.chat.id, f'Олимпиады под номером {message.text} нет в списке')
            button_menu_global(message)

    def pick_lesson(message):
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        list_of_subjects = []
        schet = 1
        for i in lessons_dict:
            list_of_subjects.append(
                types.InlineKeyboardButton(text=f"‍👨‍🏫{lessons_dict[i]}", callback_data=f'lesson{schet}'))
            schet += 1
        keyboard.add(*list_of_subjects)
        olympiad_name = main_dict[message.chat.id][1]
        bot.send_message(message.chat.id, f'📌Выбери предмет, по которому пишешь {olympiad_name}', reply_markup=keyboard)

    def pick_participation_stage(message):
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        passed_registration = types.InlineKeyboardButton(text="✏️Прошёл регистрацию",
                                                         callback_data='passed_registration')
        wrote_qualifying = types.InlineKeyboardButton(text="📝Написал отборочный этап", callback_data='wrote_qualifying')
        passed_final = types.InlineKeyboardButton(text="🔝Прошёл на заключительный этап", callback_data='passed_final')
        took_final = types.InlineKeyboardButton(text="🏅Принял участие в финале(участник)", callback_data='took_final')
        final_prize_winner = types.InlineKeyboardButton(text="🥈🥉Призёр финала(диплом 2 или 3 степени)",
                                                        callback_data='final_prize_winner')
        winner_of_final = types.InlineKeyboardButton(text="🥇Победитель финала(диплом 1 степени)",
                                                     callback_data='winner_of_final')

        keyboard.add(passed_registration, wrote_qualifying, passed_final, took_final, final_prize_winner,
                     winner_of_final)

        olympiad_name = main_dict[message.chat.id][1]
        bot.send_message(message.chat.id,
                         "☑️Вы выбрали олимпиаду: " + olympiad_name +
                         '\n 📚Укажите этап участия в олимпиаде в данный момент', reply_markup=keyboard)

    def pick_mentor(message):

        keyboard = types.InlineKeyboardMarkup(row_width=1)
        schet = 1
        list_of_mentors = []
        for i in mentors_dict:
            list_of_mentors.append(
                types.InlineKeyboardButton(text=f"‍👨‍🏫{mentors_dict[i]}", callback_data=f'mentor{schet}'))
            schet += 1
        keyboard.add(*list_of_mentors)
        bot.send_message(message.chat.id, '📌Выбери своего наставника', reply_markup=keyboard)

    def correct_chak_olymp(message):
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        yes_button = types.InlineKeyboardButton(text="Да", callback_data='correct')
        no_button = types.InlineKeyboardButton(text="Нет", callback_data='not_correct')
        keyboard.add(yes_button, no_button)

        name_user = main_dict[message.chat.id][0]
        olympiad_name = main_dict[message.chat.id][1]

        if 'НТО' in olympiad_name:
            part_stage = main_dict[message.chat.id][2]
            mentor = main_dict[message.chat.id][3]
            bot.send_message(message.chat.id,
                             f'👦Имя: {name_user}\n'
                             f'🏆Олимпиада: {olympiad_name}\n'
                             f'📚Предмет: -\n'
                             f'📎Этап: {part_stage}\n'
                             f"📌Наставник: {mentor}", reply_markup=keyboard)
        else:
            lesson = main_dict[message.chat.id][2]
            part_stage = main_dict[message.chat.id][3]
            mentor = main_dict[message.chat.id][4]
            bot.send_message(message.chat.id,
                             f'👦Имя: {name_user}\n'
                             f'🏆Олимпиада: {olympiad_name}\n'
                             f'📚Предмет: {lesson}\n'
                             f'📎Этап: {part_stage}\n'
                             f"📌Наставник: {mentor}", reply_markup=keyboard)

    def new_olymp_reg(message):
        name_user = main_dict[message.chat.id][0]
        olympiad_name = main_dict[message.chat.id][1]

        clas = cursor.execute(f"SELECT class FROM user"
                              f" WHERE user_name = '{name_user}'").fetchall()[0][0]

        if 'НТО' in olympiad_name:
            part_stage = main_dict[message.chat.id][2]
            mentor = main_dict[message.chat.id][3]
            cursor_olymp.execute('INSERT INTO olympiad VALUES (?, ?, ?, ?, ?, ?, ?)',
                                 (None, name_user, olympiad_name, '-', part_stage, mentor, clas))
        else:
            lesson = main_dict[message.chat.id][2]
            part_stage = main_dict[message.chat.id][3]
            mentor = main_dict[message.chat.id][4]
            cursor_olymp.execute('INSERT INTO olympiad VALUES (?, ?, ?, ?, ?, ?, ?)',
                                 (None, name_user, olympiad_name, lesson, part_stage, mentor, clas))
        db_olympiad.commit()

    def user_olympiad_list(message):
        user_name = receive_user_name(message)
        cursor_olymp.execute("SELECT * FROM olympiad")
        massive_big = cursor_olymp.fetchall()
        user_olympiad_string = '<b>📄Список ваших олимпиад:</b>\n\n'
        schet = 1
        for i in range(len(massive_big)):
            if massive_big[i][1] == user_name:
                if 'НТО' in massive_big[i][2]:
                    user_olympiad_string += f'<b>{schet})🎯Олимпиада:</b> {massive_big[i][2]}\n' \
                                            f'<b>Предмет:</b> -\n' \
                                            f'<b>Этап:</b> {massive_big[i][4]}\n' \
                                            f'<b>Наставник:</b> {massive_big[i][5]}\n' \
                                            f'<b>Индекс:</b> {massive_big[i][0]}\n\n'
                else:
                    user_olympiad_string += f'<b>{schet}) 🎯Олимпиада:</b> {massive_big[i][2]}\n' \
                                            f'<b>Предмет:</b> {massive_big[i][3]}\n' \
                                            f'<b>Этап:</b> {massive_big[i][4]}\n' \
                                            f'<b>Наставник:</b> {massive_big[i][5]}\n' \
                                            f'<b>Индекс:</b> {massive_big[i][0]}\n\n'
                schet += 1

        global olympiad_list_emptily
        olympiad_list_emptily = False

        if len(user_olympiad_string) == 32:
            olympiad_list_emptily = True
            user_olympiad_string += 'Похоже список пуст'
            bot.send_message(message.chat.id, user_olympiad_string, parse_mode="HTML")
        else:
            bot.send_message(message.chat.id, user_olympiad_string, parse_mode="HTML")

    def changestage(message):
        keyboard = types.InlineKeyboardMarkup(row_width=1)
        passed_registration = types.InlineKeyboardButton(text="✏️Прошёл регистрацию",
                                                         callback_data='reg')
        wrote_qualifying = types.InlineKeyboardButton(text="📝Написал отборочный этап", callback_data='wrote')
        passed_final = types.InlineKeyboardButton(text="🔝Прошёл на заключительный этап", callback_data='final')
        took_final = types.InlineKeyboardButton(text="🏅Принял участие в финале(участник)", callback_data='lookfinal')
        final_prize_winner = types.InlineKeyboardButton(text="🥈🥉Призёр финала(диплом 2 или 3 степени)",
                                                        callback_data='prizer')
        winner_of_final = types.InlineKeyboardButton(text="🥇Победитель финала(диплом 1 степени)",
                                                     callback_data='winner')

        keyboard.add(passed_registration, wrote_qualifying, passed_final, took_final, final_prize_winner,
                     winner_of_final)

        bot.send_message(message.chat.id,
                         '📚Укажите этап участия в олимпиадах, по которому хотите получить рейтинг.',
                         reply_markup=keyboard)

    def diagrama(message, groups, counts, Title, save):
        mx = len(max(groups, key=len))
        plt.figure(figsize=(5 + mx // 7, 8))
        obj = plt.barh(groups, counts, edgecolor='black', linewidth=0.4)

        for i, v in enumerate(counts):
            plt.text(v + 0.1, i, str(v), color='black', fontweight='bold', size=12)

        ax = plt.gca()
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_visible(False)
        plt.tick_params(bottom=False)
        plt.xticks([])

        plt.xlabel("Количество написанных олимпиад", fontweight='bold')
        plt.savefig(save)
        bot.send_document(message.chat.id, open(save, "rb"))
        ax.remove()
        obj.remove()

    def my_rating_place(message, stage):  # эта функция отвечает за статистику
        matplotlib.rcParams['figure.subplot.left'] = 0.5

        stage = participation_stage_dict[stage]
        olympiad_counter_list = cursor_olymp.execute(f"SELECT * FROM olympiad").fetchall()
        clas_raiting = {}
        predmet_raiting = {}
        user_raiting = {}
        olymp_raiting = {}
        for i in olympiad_counter_list:
            if i[4] == stage:
                if i[-1] in clas_raiting:
                    clas_raiting[i[-1]] += 1
                else:
                    clas_raiting[i[-1]] = 1

                if i[3] in predmet_raiting:
                    predmet_raiting[i[3]] += 1
                else:
                    predmet_raiting[i[3]] = 1

                if i[1] in user_raiting:
                    user_raiting[i[1]] += 1
                else:
                    user_raiting[i[1]] = 1

                if i[2] in olymp_raiting:
                    olymp_raiting[i[2]] += 1
                else:
                    olymp_raiting[i[2]] = 1

        groups = []
        counts = []
        flag = 0
        if len(user_raiting) > 0:
            flag = 1
            spisok = []
            for i in user_raiting:
                spisok.append((i, user_raiting[i]))
            spisok.sort(key=lambda x: x[1])

            for i in spisok:
                groups.append(i[0])
                counts.append(i[1])

            diagrama(message, groups, counts, "", "raiting_of_users.png")
            groups.clear()
            counts.clear()
        if len(olymp_raiting) > 0:
            flag = 1
            flag = 1
            spisok = []
            for i in olymp_raiting:
                spisok.append((i, olymp_raiting[i]))
            spisok.sort(key=lambda x: x[1])

            for i in spisok:
                groups.append(i[0])
                counts.append(i[1])
            diagrama(message, groups, counts, "", "raiting_of_olymp.png")
            groups.clear()
            counts.clear()

        if len(clas_raiting) > 0:
            flag = 1
            spisok = []
            for i in clas_raiting:
                spisok.append((i, clas_raiting[i]))
            spisok.sort(key=lambda x: x[1])

            for i in spisok:
                groups.append(i[0])
                counts.append(i[1])
            diagrama(message, groups, counts, "", "raiting_of_claces.png")
            groups.clear()
            counts.clear()
        if len(predmet_raiting) > 0:
            flag = 1
            spisok = []
            for i in predmet_raiting:
                if i != "-":
                    spisok.append((i, predmet_raiting[i]))
                else:
                    spisok.append(("НТО", predmet_raiting[i]))
            spisok.sort(key=lambda x: x[1])

            for i in spisok:
                groups.append(i[0])
                counts.append(i[1])
            if len(groups) > 0:
                flag = 1
                diagrama(message, groups, counts, f"Рейтинг предметов по параметру: '{stage}'", "raiting_of_lesson.png")
                groups.clear()
                counts.clear()
        if not flag:
            bot.send_message(message.chat.id, "Нет данных")

    def olympiad_rating(message):
        olympiad_counter_list = cursor_olymp.execute(f"SELECT * FROM olympiad_rating").fetchall()
        olympiad_counter_dict = {}

        for i in range(len(olympiad_counter_list)):
            olympiad_counter_dict[olympiad_counter_list[i][1]] = olympiad_counter_list[i][2]

        olympiad_counter_dict = {k: v for k, v in
                                 sorted(olympiad_counter_dict.items(), key=lambda item: item[1], reverse=True)}
        rating = ''
        schet = 1
        for k, v in olympiad_counter_dict.items():
            rating += f'<b>{schet}) {k}\n</b>' \
                      f'Количество олимпиад: {v}\n'
            schet += 1

        bot.send_message(message.chat.id, rating, parse_mode='HTML')
        del olympiad_counter_dict
        button_message(message)

    def rename_user(message):
        past_user_name = receive_user_name(message)

        clases = ["6a", "6b", "6c", "7a", "7b", "7c", "8a", "8b", "8c", "9a", "9b", "9c", "10a", "10b", "10c", "11a",
                  "11b",
                  "11c"]

        global new_user_name
        global class_name

        try:
            text = message.text.split(",")
            new_user_name = text[0].lstrip().rstrip()
            class_name = text[1].lower().lstrip().rstrip()

            if class_name not in clases:
                bot.send_message(message.chat.id, "Ошибка ввода класса!")
                button_menu_profile(message)
            else:
                spisok = cursor.execute(
                    f"SELECT class FROM user WHERE user_name = '{new_user_name}'").fetchone()

                cursor.execute(
                    f"SELECT user_name FROM user WHERE user_name = '{new_user_name}'")  # Проверяем, есть ли написанное имя в бд
                if cursor.fetchone() is None or (past_user_name == new_user_name):  # Если его нет, тогда идём дальше
                    keyboard = types.InlineKeyboardMarkup(row_width=2)
                    yes_button = types.InlineKeyboardButton(text="Да", callback_data='yes_rename')
                    no_button = types.InlineKeyboardButton(text="Нет", callback_data='no_not_rename')
                    keyboard.add(yes_button, no_button)

                    bot.send_message(message.chat.id,
                                     f'‼️ВНИМАНИЕ‼️\nЕсли вы смените своё ФИО, все олимпиады, которые были записаны будут записаны '
                                     f'на новое введённое ФИО\n\nТекущее ФИО:{past_user_name}\nНовое ФИО: {new_user_name}\n'
                                     f'Ваш текущий класс: {class_name}\n\n'
                                     f'Вы уверены, что хотите поменять текущее имя/класс на новое?',
                                     reply_markup=keyboard)
                else:
                    bot.send_message(message.chat.id,
                                     'Такое имя пользователя c таким классом уже зарегистрировано. Если вы хотите поменять ваше ФИО на другое,'
                                     'то попробуйте заново')
                    button_menu_profile(message)


        except Exception:
            bot.send_message(message.chat.id, "Ошибка ввода!")
            button_menu_profile(message)

    def end_rename_user(message):
        past_user_name = receive_user_name(message)

        result_user = cursor.execute(f'SELECT id FROM user WHERE user_name="{past_user_name}"').fetchall()
        result_olympiad = cursor_olymp.execute(f'SELECT id FROM olympiad WHERE user_name="{past_user_name}"').fetchall()

        for user_id in result_user:  # перебираем результаты
            cursor.execute('UPDATE user SET user_name=?, class=? WHERE id=?',
                           (new_user_name, class_name, user_id[0]))
            # обновляем записи в таблице user
            db.commit()

        for olympiad_id in result_olympiad:
            cursor_olymp.execute('UPDATE olympiad SET user_name=?, class=? WHERE id=?',
                                 (new_user_name, class_name, olympiad_id[0]))  # обновляем записи в таблице.
            db_olympiad.commit()

        bot.send_message(message.chat.id, 'Ваше ФИО и/или класс обновлены', reply_markup=markup)
        cursor_olymp.execute(
            'UPDATE olympiad_rating SET user_name=? WHERE user_name=?',
            (new_user_name, past_user_name))

        db_olympiad.commit()
        button_menu_profile(message)

    def del_olympiad(message):
        name_user = receive_user_name(message)
        len_do = cursor_olymp.execute('SELECT olympiad FROM olympiad WHERE user_name=?', (name_user,)).fetchall()
        cursor_olymp.execute("DELETE FROM olympiad WHERE id=?", (message.text,))
        db_olympiad.commit()

        len_posle = cursor_olymp.execute('SELECT olympiad FROM olympiad WHERE user_name=?', (name_user,)).fetchall()
        if len_do != len_posle:
            cursor_olymp.execute(
                'UPDATE olympiad_rating SET olympiad_counter=olympiad_counter - ? WHERE user_name=?',
                (1, name_user))
            db_olympiad.commit()
            bot.send_message(message.chat.id, f'Олимпиада под номером {message.text} успешно удалена')
        else:
            bot.send_message(message.chat.id, f'Олимпиады под индексом {message.text} нет в списке ваших олимпиад')
        button_menu_profile(message)

    def tablitsa(message):
        workbook = xlsxwriter.Workbook('olympiads.xlsx')
        worksheet = workbook.add_worksheet()
        data = cursor_olymp.execute(
            'SELECT user_name, olympiad, lesson, participation_stage, mentor, class FROM olympiad').fetchall()

        for row, (user_name, olympiad, lesson, participation_stage, mentor, clas) in enumerate(data):
            worksheet.write(row, 0, user_name)
            worksheet.write(row, 1, olympiad)
            worksheet.write(row, 2, lesson)
            worksheet.write(row, 3, participation_stage)
            worksheet.write(row, 4, mentor)
            worksheet.write(row, 5, clas)
        workbook.close()
        f = open('olympiads.xlsx', "rb")
        bot.send_document(message.chat.id, f)

    @bot.callback_query_handler(func=lambda call: True)
    def answer(call):

        cursor.execute(f"SELECT olympiad_index FROM user WHERE chat_id = '{call.message.chat.id}'")
        olympiad_index = int(cursor.fetchone()[0])

        if call.data == 'next_15':
            bot.delete_message(call.message.chat.id, stickerId.id)
            bot.delete_message(call.message.chat.id, call.message.id)
            olympiad_index += 1
            cursor.execute(
                f"UPDATE user set olympiad_index = '{olympiad_index}' WHERE chat_id = '{call.message.chat.id}'")
            send_olympiad_list(call.message, olympiad_index)

        elif call.data == 'past_15':
            bot.delete_message(call.message.chat.id, stickerId.id)
            bot.delete_message(call.message.chat.id, call.message.id)
            olympiad_index -= 1
            cursor.execute(
                f"UPDATE user set olympiad_index = '{olympiad_index}' WHERE chat_id = '{call.message.chat.id}'")
            send_olympiad_list(call.message, olympiad_index)

        elif call.data == 'pick_olympiad':

            msg = bot.send_message(call.message.chat.id,
                                   '📋Выбери номер нужной тебе олимпиады\n\nЕсли ты выбираешь какой то из профилей НТО, '
                                   'не забудь указать через нижнее подчёркивание второе число')
            bot.register_next_step_handler(msg, pick_olymp)

        elif 'lesson' in call.data:
            name_user = receive_user_name(call.message)
            olympiad_name = main_dict[call.message.chat.id][1]
            lesson = lessons_dict[call.data]
            main_dict[call.message.chat.id].append(lesson)
            bot.delete_message(call.message.chat.id, call.message.id)
            pick_participation_stage(call.message)


        elif call.data in (
                'passed_registration', 'wrote_qualifying', 'passed_final', 'took_final', 'final_prize_winner',
                'winner_of_final'):
            name_user = main_dict[call.message.chat.id][0]
            olympiad_name = main_dict[call.message.chat.id][1]
            lesson = main_dict[call.message.chat.id][-1]
            all_olympiad = cursor_olymp.execute(f"SELECT * FROM olympiad WHERE user_name='{name_user}'").fetchall()
            flag = False

            for i in range(len(all_olympiad)):
                if 'НТО' in olympiad_name:
                    if all_olympiad[i][2] == olympiad_name and all_olympiad[i][4] == participation_stage_dict[call.data]:
                        flag = True
                else:
                    if all_olympiad[i][2] == olympiad_name and all_olympiad[i][3] == lesson and all_olympiad[i][4] == \
                            participation_stage_dict[call.data]:
                        flag = True
            if not flag:
                main_dict[call.message.chat.id].append(participation_stage_dict[call.data])
                bot.delete_message(call.message.chat.id, call.message.id)
                pick_mentor(call.message)
            else:
                bot.send_message(call.message.chat.id,
                                 f'Вы уже регистрировали {olympiad_name} на данной стадии участия\n\n')
                button_menu_global(call.message)

        elif call.data in (
                'reg', 'wrote', 'final', 'lookfinal', 'prizer',
                'winner'):
            data = call.data

            if data == "reg":
                data = 'passed_registration'
            elif data == "wrote":
                data = "wrote_qualifying"
            elif data == 'final':
                data = "passed_final"
            elif data == "lookfinal":
                data = "took_final"
            elif data == "prizer":
                data = 'final_prize_winner'
            else:
                data = 'winner_of_final'

            my_rating_place(call.message, data)

        elif 'mentor' in call.data:
            main_dict[call.message.chat.id].append(mentors_dict[call.data])

            bot.delete_message(call.message.chat.id, call.message.id)
            correct_chak_olymp(call.message)

        elif call.data in ('correct', 'not_correct'):
            if call.data == 'correct':
                try:
                    new_olymp_reg(call.message)

                    bot.delete_message(call.message.chat.id, call.message.id)
                    bot.send_message(call.message.chat.id, 'Вы успешно зарегистрировали олимпиаду')
                    lessons = []
                    name_user = receive_user_name(call.message)
                    cursor_olymp.execute(
                        'UPDATE olympiad_rating SET olympiad_counter=olympiad_counter + ? WHERE user_name=?',
                        (1, name_user))
                    db_olympiad.commit()
                    del main_dict[call.message.chat.id]
                    button_menu_global(call.message)
                except EOFError:
                    bot.send_message(call.message.chat.id, f'При регистрации олимпиады произошла ошибка')
            else:
                bot.send_message(call.message.chat.id, 'Олимпиада не зарегистрирована')
                button_menu_global(call.message)

        elif call.data in ('yes_rename', 'no_not_rename'):
            if call.data == 'yes_rename':
                end_rename_user(call.message)
            else:
                bot.send_message(call.message.chat.id, 'Имя пользователя не сменилось')
                button_menu_profile(call.message)

    @bot.message_handler(content_types=['document'])
    def handle_docs_photo(message):
        try:

            name = receive_user_name(message)

            if name == "Администратор":

                file_info = bot.get_file(message.document.file_id)
                downloaded_file = bot.download_file(file_info.file_path)

                src = "peregruz.xlsx"
                with open(src, 'wb') as new_file:
                    new_file.write(downloaded_file)

                spisok = []
                slovar_schet = {}
                # Define variable to load the wookbook
                wookbook = openpyxl.load_workbook(src)
                # Define variable to read the active sheet:
                worksheet = wookbook.active
                cursor_olymp.execute("DELETE FROM olympiad")
                users = cursor.execute("SELECT user_name FROM user").fetchone()

                anonim = []
                for i in range(0, worksheet.max_row):
                    podspisok = [None]
                    for col in worksheet.iter_cols(1, worksheet.max_column):
                        podspisok.append(col[i].value)

                    if podspisok[1] in users:
                        if podspisok[1] in slovar_schet:
                            spisok.append(podspisok)
                            slovar_schet[podspisok[1]] += 1
                        else:
                            slovar_schet[podspisok[1]] = 1
                    else:
                        anonim.append(podspisok[1])
                cursor_olymp.execute("DELETE FROM olympiad_rating")

                for i in users:
                    if i not in slovar_schet:
                        slovar_schet[i] = 0

                for i in slovar_schet:
                    cursor_olymp.execute(f'INSERT INTO olympiad_rating VALUES(?,?,?);', (None, i, slovar_schet[i]))

                cursor_olymp.executemany('INSERT INTO olympiad VALUES(?,?,?, ?,?,?, ?);', spisok)
                db_olympiad.commit()

                if len(anonim) > 0:
                    bot.send_message(message.chat.id, f"{anonim} \n не зарегестрированы и не были добавлены.")
                bot.send_message(message.chat.id, "База данных обновлена")

            else:
                bot.send_message(message.chat.id, "Отказано в доступе")

        except Exception as e:
            bot.reply_to(message, "")

    @bot.message_handler(content_types=['text'])
    def button_answer(message):
        text = message.text

        cursor.execute(f"SELECT chat_id FROM user WHERE chat_id = '{message.chat.id}'")
        if cursor.fetchone() is None:
            if 'reg' in message.text.split(', '):
                name(message)

        elif text == '👤Мой профиль':
            button_menu_profile(message)

        elif text == '📋РсОШ трекер':
            button_menu_global(message)

        elif text == '📊Получить данные':
            button_menu_data(message)

        elif text == '🔙Назад':
            button_message(message)

        elif text == '✏️Зарегистрироваться на олимпиаду':
            global ind
            ind = 0
            cursor.execute(
                f"UPDATE user set olympiad_index = '{0}' WHERE chat_id = '{message.chat.id}'")
            send_olympiad_list(message, 0)

        elif text == "🗑Удалить олимпиаду":
            user_olympiad_list(message)
            global olympiad_list_emptily
            if olympiad_list_emptily:
                bot.send_message(message.chat.id, 'Тут нечего удалять')
                button_menu_profile(message)
            else:
                msg = bot.send_message(message.chat.id, 'Напишите ИНДЕКС олимпиады, которую хотите удалить',
                                       reply_markup=types.ReplyKeyboardRemove())
                bot.register_next_step_handler(msg, del_olympiad)
        elif text == '📄Мои олимпиады':
            user_olympiad_list(message)
            button_menu_profile(message)
        elif text == '📊Статистика олимпиад':
            changestage(message)
        elif text == '🔃Поменять имя и/или клас':
            msg = bot.send_message(message.chat.id, 'Введите новое ФИО и класс через запятую\n\n'
                                                    'Пример: Иванов Иван Иванович, 11A',
                                   reply_markup=types.ReplyKeyboardRemove())
            bot.register_next_step_handler(msg, rename_user)
        elif text == '🏆Рейтинг':
            olympiad_rating(message)
        elif text == "👀Получить Excel-таблицу":
            tablitsa(message)

    while True:
        try:
            bot.polling(none_stop=True)
        except Exception as _ex:
            print(_ex)
            sleep(15)


if __name__ == '__main__':
    main()
